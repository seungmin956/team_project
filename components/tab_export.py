# components/tab_export.py - Excel 리포트(Tableau 포함) 최종 개선 버전

import streamlit as st
import pandas as pd
from datetime import datetime, timedelta 
import json
import os
from functools import lru_cache

# openpyxl 관련 import
import openpyxl
from openpyxl.drawing import image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import requests
import io
from PIL import Image

# LangChain imports
from langchain_openai import ChatOpenAI
from langchain.schema import HumanMessage
from dotenv import load_dotenv

try:
    from chart_downloader import add_auto_download_button, download_all_tableau_charts
    SELENIUM_AVAILABLE = True
except ImportError:
    SELENIUM_AVAILABLE = False

import streamlit.components.v1 as components

load_dotenv()

# =============================================================================
# 캐시된 프로젝트 로딩
# =============================================================================

@st.cache_data(ttl=300)  # 5분 TTL
def _load_all_histories():
    """모든 대화 기록을 캐시와 함께 로드"""
    try:
        if not os.path.exists("chat_histories.json"):
            return {}
        with open("chat_histories.json", 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        st.error(f"파일 로드 실패: {e}")
        return {}

# =============================================================================
# 메인 함수
# =============================================================================

def show_export_helper():
    """수출 제안서 도우미 메인 함수"""

    # 안내 메시지 
    st.info("""
    📝 **수출용 상품 기획안 작성 도우미**

    - 챗봇 질의응답 시 설정한 프로젝트명 선택  
    - 제품 정보, 제안 의도 입력 시 EXCEL 파일 생성 기능 제공  
    - TAB 1의 그래프, 챗봇 질의응답 데이터 함께 출력 가능
    """)
    
    # 세션 상태 초기화
    init_session_state()

    col1, col2 = st.columns([1, 1])
    
    with col1:
        st.markdown("**1️⃣ 정보 입력**")
        show_basic_info_form()
    
    with col2:
        st.markdown("**2️⃣ 예시 이미지**")
        render_guide_section()

# =============================================================================
# 세션 상태 및 기본 함수들
# =============================================================================

def init_session_state():
    """세션 상태 초기화"""
    defaults = {
        "export_data": {},
        "report_generated": False,
        "selected_template_data": "기본 제안서",
        "show_summary_area": False,
        "summary_content": "",
        "ai_processing": False
    }
    
    for key, default_value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = default_value

# =============================================================================
# 자동 다운로드 함수들
# =============================================================================

def auto_download_all_tableau_charts():
    """Tableau 차트 4개 자동 다운로드"""
    
    chart_configs = {
        "state_food.png": {
            "url": "https://public.tableau.com/static/images/st/state_food_exp2_17479635670940/State/1.png",
            "title": "🗺️ 미국 주별 식품 지출"
        },
        "food_trend.png": {
            "url": "https://public.tableau.com/static/images/ma/main01/1_1/1.png",
            "title": "📈 연도별 식품 지출 추이"
        },
        "recall_heatmap.png": {
            "url": "https://public.tableau.com/static/images/fo/food_recall_year_01/1_1/1.png", 
            "title": "🔥 리콜 원인별 히트맵"
        },
        "recall_class.png": {
            "url": "https://public.tableau.com/static/images/fo/food_recall_class_01/1_1/1.png",
            "title": "📊 리콜 등급별 발생 건수"
        }
    }
    
    # charts 폴더 생성
    charts_dir = "./charts"
    if not os.path.exists(charts_dir):
        os.makedirs(charts_dir)
    
    success_count = 0
    failed_downloads = []
    
    for filename, config in chart_configs.items():
        try:
            st.write(f"⏬ {config['title']} 다운로드 중...")
            
            # 이미지 다운로드
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
            }
            
            response = requests.get(config['url'], headers=headers, timeout=15)
            response.raise_for_status()
            
            # 이미지 처리
            if response.content and 'image' in response.headers.get('content-type', ''):
                pil_image = Image.open(io.BytesIO(response.content))
                
                # 적절한 크기로 조정
                pil_image.thumbnail((800, 600), Image.Resampling.LANCZOS)
                
                # 파일 저장
                file_path = os.path.join(charts_dir, filename)
                pil_image.save(file_path, "PNG")
                
                if os.path.exists(file_path) and os.path.getsize(file_path) > 0:
                    success_count += 1
                    st.success(f"✅ {filename} 저장 완료")
                else:
                    failed_downloads.append(filename)
                    st.error(f"❌ {filename} 저장 실패")
            else:
                failed_downloads.append(filename)
                st.error(f"❌ {filename} - 유효하지 않은 이미지")
                
        except Exception as e:
            failed_downloads.append(filename)
            st.error(f"❌ {filename} 다운로드 실패: {str(e)[:50]}...")
    
    return success_count, failed_downloads

def render_auto_download_section():
    """자동 다운로드 섹션 렌더링"""
        
    # 현재 저장된 차트 상태 확인
    chart_files = [
        "./charts/state_food.png",
        "./charts/food_trend.png", 
        "./charts/recall_heatmap.png",
        "./charts/recall_class.png"
    ]
    
    existing_count = sum(1 for f in chart_files if os.path.exists(f))
    
    if existing_count > 0:
        st.info(f"📊 현재 저장된 차트: {existing_count}/4개")
    
    # 자동 다운로드 버튼
    if st.button("🚀 Tab 1 차트 이미지 자동 다운로드", use_container_width=True):
        with st.spinner("📥 모든 차트를 자동으로 다운로드하는 중..."):
            success_count, failed_downloads = auto_download_all_tableau_charts()
            
            st.markdown("---")
            
            if success_count == 4:
                st.success("🎉 모든 차트 다운로드 완료!")
            elif success_count > 0:
                st.warning(f"⚠️ {success_count}/4개 차트만 다운로드됨")
                if failed_downloads:
                    st.error(f"실패한 파일: {', '.join(failed_downloads)}")
            else:
                st.error("❌ 모든 다운로드 실패. 네트워크 연결을 확인해주세요.")
    
    st.markdown("---")

def render_guide_section():
    """가이드 이미지 및 차트 캡처 섹션"""
    try:
        st.image('./가이드.png')
    except FileNotFoundError:
        st.warning("이미지 파일을 찾을 수 없습니다.")
    except Exception as e:
        st.error(f"이미지 로드 오류: {e}")
    
    st.markdown("---")
    
    # 자동 다운로드 섹션
    render_auto_download_section()
    
    # # 저장된 차트 확인
    # if st.button("📂 저장된 차트 확인", use_container_width=True):
    #     check_saved_charts()
    
    # st.markdown("---")
    
    # Excel 생성 버튼
    add_single_excel_button()

def show_streamlit_chart_capture():
    """Streamlit에서 차트 캡처하기"""
    
    st.markdown("### 📊 차트 저장하기")
    st.info("""
    **📝 저장 방법:**
    1. 아래 각 차트에서 우클릭 → "이미지로 저장" 선택
    2. 파일명을 정확히 입력: `state_food.png`, `food_trend.png`, `recall_heatmap.png`, `recall_class.png`
    3. `charts` 폴더에 저장 (없으면 자동 생성됨)
    """)
    
    # charts 폴더 생성
    create_charts_folder()
    
    # 차트별 탭
    tab1, tab2, tab3, tab4 = st.tabs(["🗺️ 주별 식품 지출", "📈 연도별 추이", "🔥 리콜 히트맵", "📊 리콜 등급별"])
    
    with tab1:
        st.markdown("**🗺️ 미국 주별 식품 지출** - 저장명: `state_food.png`")
        
        components.html("""
        <div class='tableauPlaceholder' id='viz_state_food' style='position:relative; width:100%; height:600px;'>
        <noscript>
            <a href='#'>
            <img src='https://public.tableau.com/static/images/st/state_food_exp2_17479635670940/State/1.png'
                style='border:none' alt='State'>
            </a>
        </noscript>
        <object class='tableauViz' style='display:none;'>
            <param name='host_url' value='https%3A%2F%2Fpublic.tableau.com%2F' />
            <param name='embed_code_version' value='3' />
            <param name='site_root' value='' />
            <param name='name' value='state_food_exp2_17479635670940/State' />
            <param name='tabs' value='no' />
            <param name='toolbar' value='bottom' />
            <param name='language' value='ko-KR' />
        </object>
        </div>
        <script src='https://public.tableau.com/javascripts/api/viz_v1.js'></script>
        <script>
        const div = document.getElementById("viz_state_food");
        const vizOB = div.getElementsByTagName("object")[0];
        function resizeViz(){
            vizOB.style.width = "100%";
            vizOB.style.height = "600px";
        }
        resizeViz();
        window.addEventListener("resize", resizeViz);
        </script>
        """, height=650)
        
        st.success("💾 **저장**: 차트 우클릭 → '이미지로 저장' → `state_food.png`")
    
    with tab2:
        st.markdown("**📈 연도별 식품 지출 추이** - 저장명: `food_trend.png`")
        
        components.html("""
        <div class='tableauPlaceholder' id='viz_food_trend' style='position:relative; width:100%; height:600px;'>
        <noscript>
            <a href='#'>
            <img src='https://public.tableau.com/static/images/ma/main01/1_1/1.png'
                style='border:none' alt='대시보드 1'>
            </a>
        </noscript>
        <object class='tableauViz' style='display:none;'>
            <param name='host_url' value='https%3A%2F%2Fpublic.tableau.com%2F'/>
            <param name='embed_code_version' value='3'/>
            <param name='site_root' value=''/>
            <param name='name' value='main01/1_1'/>
            <param name='tabs' value='no'/>
            <param name='toolbar' value='yes'/>
            <param name='language' value='ko-KR'/>
        </object>
        </div>
        <script src='https://public.tableau.com/javascripts/api/viz_v1.js'></script>
        <script>
        const divEl = document.getElementById('viz_food_trend');
        const vizEl = divEl.getElementsByTagName('object')[0];
        vizEl.style.width = '100%';
        vizEl.style.height = '600px';
        </script>
        """, height=650)
        
        st.success("💾 **저장**: 차트 우클릭 → '이미지로 저장' → `food_trend.png`")
    
    with tab3:
        st.markdown("**🔥 리콜 원인별 히트맵** - 저장명: `recall_heatmap.png`")
        
        components.html("""
        <div class='tableauPlaceholder' id='viz_recall_heatmap' style='position:relative; width:100%; height:600px;'>
        <noscript>
            <a href='#'>
            <img src='https://public.tableau.com/static/images/fo/food_recall_year_01/1_1/1_rss.png'
                style='border:none' alt='연도별 리콜건수 변화 히트맵'>
            </a>
        </noscript>
        <object class='tableauViz' style='display:none;'>
            <param name='host_url' value='https%3A%2F%2Fpublic.tableau.com%2F'/>
            <param name='embed_code_version' value='3'/>
            <param name='site_root' value=''/>
            <param name='name' value='food_recall_year_01/1_1'/>
            <param name='tabs' value='no'/>
            <param name='toolbar' value='yes'/>
            <param name='language' value='ko-KR'/>
        </object>
        </div>
        <script src='https://public.tableau.com/javascripts/api/viz_v1.js'></script>
        <script>
        const divElement = document.getElementById('viz_recall_heatmap');
        const vizElement = divElement.getElementsByTagName('object')[0];
        vizElement.style.width = '100%';
        vizElement.style.height = '600px';
        </script>
        """, height=650)
        
        st.success("💾 **저장**: 차트 우클릭 → '이미지로 저장' → `recall_heatmap.png`")
    
    with tab4:
        st.markdown("**📊 리콜 등급별 발생 건수** - 저장명: `recall_class.png`")
        
        components.html("""
        <div class='tableauPlaceholder' id='viz_recall_class' style='position:relative; width:100%; height:600px;'>
        <noscript>
            <a href='#'>
            <img alt='리콜 등급(Class)별 발생 건수' src='https://public.tableau.com/static/images/fo/food_recall_class_01/1_1/1_rss.png' style='border: none' />
            </a>
        </noscript>
        <object class='tableauViz' style='display:none;'>
            <param name='host_url' value='https%3A%2F%2Fpublic.tableau.com%2F' />
            <param name='embed_code_version' value='3' />
            <param name='site_root' value='' />
            <param name='name' value='food_recall_class_01/1_1' />
            <param name='tabs' value='no' />
            <param name='toolbar' value='yes' />
            <param name='language' value='ko-KR' />
        </object>
        </div>
        <script src='https://public.tableau.com/javascripts/api/viz_v1.js'></script>
        <script>
        const divElement = document.getElementById('viz_recall_class');
        const vizElement = divElement.getElementsByTagName('object')[0];
        vizElement.style.width = '100%';
        vizElement.style.height = '600px';
        </script>
        """, height=650)
        
        st.success("💾 **저장**: 차트 우클릭 → '이미지로 저장' → `recall_class.png`")

def create_charts_folder():
    """charts 폴더 생성"""
    charts_dir = "./charts"
    if not os.path.exists(charts_dir):
        os.makedirs(charts_dir)
        st.info(f"📁 `{charts_dir}` 폴더가 생성되었습니다!")
    return charts_dir

# def check_saved_charts():
#     """저장된 차트들 확인"""
    
#     chart_files = {
#         "state_food.png": "🗺️ 미국 주별 식품 지출",
#         "food_trend.png": "📈 연도별 식품 지출 추이",
#         "recall_heatmap.png": "🔥 리콜 원인별 히트맵",
#         "recall_class.png": "📊 리콜 등급별 발생 건수"
#     }
    
#     charts_dir = create_charts_folder()
    
#     st.markdown("### 📂 저장된 차트 확인")
    
#     saved_count = 0
#     cols = st.columns(2)
    
#     for i, (filename, title) in enumerate(chart_files.items()):
#         file_path = os.path.join(charts_dir, filename)
        
#         with cols[i % 2]:
#             if os.path.exists(file_path):
#                 file_size = os.path.getsize(file_path)
#                 st.success(f"✅ {filename} ({file_size} bytes)")
                
#                 # 이미지 미리보기
#                 try:
#                     st.image(file_path, caption=title, use_column_width=True)
#                     saved_count += 1
#                 except Exception as e:
#                     st.error(f"❌ {filename} 읽기 실패: {e}")
#             else:
#                 st.error(f"❌ {filename} 없음")
    
#     # 결과 요약
#     if saved_count == len(chart_files):
#             st.success("🎉 모든 차트가 준비되었습니다! Excel 생성이 가능합니다.")
#     elif saved_count > 0:
#         st.warning(f"⚠️ {saved_count}/{len(chart_files)}개 차트만 저장됨")
#     else:
#         st.info("💡 아직 저장된 차트가 없습니다. 위의 탭에서 차트를 저장해주세요.")
    
#     return saved_count

# =============================================================================
# 프로젝트 관련 함수들
# =============================================================================

def get_available_projects():
    """저장된 프로젝트 목록을 가져오는 함수"""
    try:
        all_histories = _load_all_histories()
        
        # 프로젝트명만 추출 (모드 부분 제거)
        project_names = set()
        for project_key in all_histories.keys():
            if '_' in project_key:
                parts = project_key.rsplit('_', 1)
                if len(parts) == 2 and parts[1] in ['규제', '리콜사례']:
                    project_names.add(parts[0])
            else:
                project_names.add(project_key)
        
        return sorted(list(project_names))
    except Exception as e:
        st.error(f"프로젝트 목록 불러오기 실패: {e}")
        return []

def load_project_chat_history(project_name):
    """특정 프로젝트의 통합 채팅 히스토리 불러오기"""
    try:
        all_histories = _load_all_histories()
        
        regulation_history = []
        recall_history = []
        
        regulation_key = f"{project_name}_규제"
        if regulation_key in all_histories:
            regulation_data = all_histories[regulation_key]
            regulation_history = regulation_data.get("chat_history", [])
        
        recall_key = f"{project_name}_리콜사례"
        if recall_key in all_histories:
            recall_data = all_histories[recall_key]
            recall_history = recall_data.get("chat_history", [])
        
        combined_history = regulation_history + recall_history
        return combined_history
        
    except Exception as e:
        st.error(f"프로젝트 히스토리 불러오기 실패: {e}")
        return []

def get_project_summary_info(project_name):
    """프로젝트의 요약 정보 반환"""
    try:
        all_histories = _load_all_histories()
        
        regulation_key = f"{project_name}_규제"
        recall_key = f"{project_name}_리콜사례"
        
        info = {
            "regulation_chats": 0,
            "recall_chats": 0,
            "last_updated": None,
            "modes": []
        }
        
        if regulation_key in all_histories:
            reg_data = all_histories[regulation_key]
            info["regulation_chats"] = len(reg_data.get("chat_history", [])) // 2
            info["modes"].append("규제")
            if reg_data.get("last_updated"):
                info["last_updated"] = reg_data["last_updated"]
        
        if recall_key in all_histories:
            recall_data = all_histories[recall_key]
            info["recall_chats"] = len(recall_data.get("chat_history", [])) // 2
            info["modes"].append("리콜사례")
            if recall_data.get("last_updated"):
                if not info["last_updated"] or recall_data["last_updated"] > info["last_updated"]:
                    info["last_updated"] = recall_data["last_updated"]
        
        return info
        
    except Exception as e:
        return {"regulation_chats": 0, "recall_chats": 0, "last_updated": None, "modes": []}

# =============================================================================
# UI 렌더링 함수들
# =============================================================================

def show_basic_info_form():
    """기본 정보 입력 폼"""
    narrow_col, _ = st.columns([0.8, 0.2])

    with narrow_col:
        render_project_selector()
        st.markdown("---")
        render_product_info_section()
        render_background_section()
        render_risk_summary_section()
        render_summary_display()

def render_project_selector():
    """프로젝트 선택 섹션"""
    st.markdown("**프로젝트 선택**")
    
    available_projects = get_available_projects()
    
    if available_projects:
        selected_project = st.selectbox(
            "저장된 프로젝트에서 선택",
            ["새 프로젝트"] + available_projects,
            key="project_selector",
            help="기존 프로젝트를 선택하여 규제/리콜사례 모든 Q&A 기록을 통합하여 불러옵니다."
        )
    else:
        st.info("저장된 프로젝트가 없습니다. 채팅 탭에서 대화 후 저장해주세요.")
        selected_project = "새 프로젝트"

def render_product_info_section():
    """제품 정보 입력 섹션"""
    st.markdown("**제품 정보**")
    
    product_name = st.text_input(
        "제품명", 
        placeholder="단백질 에너지바", 
        key="product_name"
    )
    
    target_market = st.text_input(
        "타겟층", 
        placeholder="30대 여성", 
        key="target_name"
    )

def render_background_section():
    """기획 의도 입력 섹션"""
    st.markdown("**기획 의도**")
    
    placeholder_text = """상세한 시장 분석, 경쟁사 내용을 입력하세요.

예시) 미국 내 30대 여성을 중심으로 고단백 식품에 대한 수요가 크게 늘고 있으며, 2022년부터 2024년까지 단백질 간식은 연평균 9%의 성장률을 기록하고 있습니다...

"""
    
    background = st.text_area(
        "내용",
        placeholder=placeholder_text,
        height=350,
        key="background"
    )

def render_risk_summary_section():
    """규제 리스크 요약 섹션"""
    st.markdown("**리스크 요약**")
    
    selected_project = st.session_state.get("project_selector", "새 프로젝트")
    
    if selected_project != "새 프로젝트":
        project_info = get_project_summary_info(selected_project)
        total_chats = project_info['regulation_chats'] + project_info['recall_chats']
        button_text = f"'{selected_project}' 프로젝트 리스크 분석"
    else:
        button_text = "현재 세션 Q&A 내용 불러오기"

    button_disabled = st.session_state.get("ai_processing", False)
    
    if st.button(button_text, disabled=button_disabled):
        process_qa_analysis(selected_project)

# =============================================================================
# AI 분석 관련 함수들
# =============================================================================

def process_qa_analysis(selected_project):
    """리스크 분석 처리"""
    st.session_state.ai_processing = True
    st.session_state.show_summary_area = True
    
    try:
        if selected_project != "새 프로젝트":
            chat_history = load_project_chat_history(selected_project)
            if not chat_history:
                st.warning(f"'{selected_project}' 프로젝트에 대화 기록이 없습니다.")
                return
        else:
            chat_history = st.session_state.get("chat_history", [])
        
        if not chat_history:
            st.warning("⚠️ 불러올 대화 기록이 없습니다. 먼저 채팅 탭에서 대화를 진행해주세요.")
            return
        
        qa_text = generate_qa_text(chat_history)
        
        if qa_text:
            perform_ai_analysis(qa_text, selected_project)
        
    except Exception as e:
        st.error(f"❌ 분석 처리 중 오류: {e}")
    finally:
        st.session_state.ai_processing = False
        st.rerun()

def generate_qa_text(chat_history):
    """채팅 히스토리에서 Q&A 텍스트 생성"""
    qa_text = ""
    for i in range(0, len(chat_history), 2):
        if i + 1 < len(chat_history):
            question = chat_history[i]["content"]
            answer = chat_history[i + 1]["content"]
            qa_text += f"질문: {question}\n답변: {answer}\n\n"
    return qa_text

@st.cache_data(ttl=1800)
def perform_ai_analysis_cached(qa_text, openai_api_key):
    """AI 분석 수행 - 캐시 적용"""
    try:
        llm = ChatOpenAI(
            model="gpt-4o-mini", 
            temperature=0.3,
            openai_api_key=openai_api_key
        )
        
        analysis_prompt = f"""
다음 Q&A 대화들을 분석하여 규제 및 리콜사례 관련 내용을 요약해주세요.

분석 요구사항:
1. 규제 관련 내용 (FDA 규정, 법령, 허가, 등록, 라벨링 등)
2. 리콜사례 관련 내용 (제품 리콜, 회수, 안전 경고 등)

각 카테고리별로 3-4문장으로 핵심 내용을 요약하고, 해당 내용이 없는 경우 "관련 내용 없음"으로 표시해주세요.

응답 형식:
📋 **규제 관련 요약**
[규제 관련 요약 내용]

🚨 **리콜 사례 요약**
[리콜 사례 관련 요약 내용]

Q&A 내용:
{qa_text}
"""
        
        response = llm.invoke([HumanMessage(content=analysis_prompt)])
        final_summary = response.content.strip()
        
        import re
        final_summary = re.sub(r'https?://[^\s]+', '', final_summary)
        final_summary = re.sub(r'📎.*?출처:.*', '', final_summary, flags=re.DOTALL)
        
        return final_summary
        
    except Exception as e:
        return f"AI 분석 실패: {str(e)}"

def perform_ai_analysis(qa_text, selected_project):
    """AI 분석 수행"""
    with st.spinner("🤖 AI가 대화 내용을 통합 분석하고 있습니다..."):
        try:
            openai_api_key = os.getenv("OPENAI_API_KEY")
            if not openai_api_key:
                st.error("OpenAI API 키가 설정되지 않았습니다.")
                return
            
            final_summary = perform_ai_analysis_cached(qa_text, openai_api_key)
            st.session_state.summary_content = final_summary
            
            if selected_project != "새 프로젝트":
                project_info = get_project_summary_info(selected_project)
                total_chats = project_info['regulation_chats'] + project_info['recall_chats']
                st.success(f"✅ '{selected_project}' 프로젝트의 {total_chats}건 Q&A를 성공적으로 분석했습니다!")
            else:
                st.success("✅ 현재 세션의 Q&A를 성공적으로 분석했습니다!")
                
        except Exception as e:
            st.error(f"❌ AI 분석 중 오류: {e}")
            st.session_state.summary_content = f"분석 실패: {e}"

def render_summary_display():
    """요약 내용 표시"""
    if st.session_state.get("show_summary_area", False):
        st.markdown("#### 📊 통합 분석 결과")
        
        edited_summary = st.text_area(
            "📝 규제/리콜 통합 분석 요약 (편집 가능)", 
            value=st.session_state.get("summary_content", ""), 
            placeholder="Q&A 내용을 불러오면 규제/리콜사례를 통합하여 분석 요약됩니다.",
            height=400,
            key="summary_editor",
            help="AI가 규제/리콜사례 모든 대화를 통합 분석한 요약입니다. 필요시 직접 편집 가능합니다."
        )
        
        if edited_summary != st.session_state.get("summary_content", ""):
            st.session_state.summary_content = edited_summary

# =============================================================================
# Tableau 이미지 관련 함수들 (개선된 버전)
# =============================================================================

def get_tableau_image_urls():
    """통일된 Tableau Public 이미지 URL들 반환"""
    return {
        "state_food": "https://public.tableau.com/static/images/st/state_food_exp2_17479635670940/State/1.png",
        "food_trend": "https://public.tableau.com/static/images/ma/main01/1_1/1.png", 
        "recall_heatmap": "https://public.tableau.com/static/images/fo/food_recall_year_01/1_1/1.png",
        "recall_class": "https://public.tableau.com/static/images/fo/food_recall_class_01/1_1/1.png"
    }

def test_tableau_urls():
    """Tableau URL 접근 가능성 테스트"""
    urls = get_tableau_image_urls()
    results = {}
    
    for name, url in urls.items():
        try:
            response = requests.head(url, timeout=10)
            results[name] = {
                "status": response.status_code,
                "accessible": response.status_code == 200,
                "content_type": response.headers.get('content-type', 'unknown')
            }
        except Exception as e:
            results[name] = {
                "status": "error",
                "accessible": False,
                "error": str(e)
            }
    
    return results

def download_tableau_image_safe(url, chart_name, max_retries=3):
    """안전한 Tableau 이미지 다운로드 with 재시도"""
    
    for attempt in range(max_retries):
        try:
            print(f"🔄 {chart_name} 다운로드 시도 {attempt + 1}/{max_retries}")
            
            # User-Agent 헤더 추가
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
            }
            
            response = requests.get(url, timeout=15, headers=headers)
            response.raise_for_status()
            
            # 응답 내용 검증
            if not response.content:
                print(f"⚠️ {chart_name}: 빈 응답 (시도 {attempt + 1})")
                continue
                
            # Content-Type 검증
            content_type = response.headers.get('content-type', '').lower()
            if 'image' not in content_type:
                print(f"⚠️ {chart_name}: 이미지가 아닌 응답 ({content_type})")
                continue
            
            # PIL로 이미지 처리
            try:
                pil_image = Image.open(io.BytesIO(response.content))
                
                # 이미지 형식 및 크기 확인
                print(f"📊 {chart_name}: {pil_image.format} {pil_image.size}")
                
                # 크기 조정
                pil_image.thumbnail((500, 350), Image.Resampling.LANCZOS)
                
                # 임시 파일 저장
                temp_filename = f"temp_{chart_name}.png"
                pil_image.save(temp_filename, "PNG")
                
                # 파일 생성 확인
                if os.path.exists(temp_filename) and os.path.getsize(temp_filename) > 0:
                    print(f"✅ {chart_name} 다운로드 성공: {temp_filename}")
                    return temp_filename
                else:
                    print(f"❌ {chart_name}: 파일 생성 실패")
                    
            except Exception as img_error:
                print(f"❌ {chart_name} 이미지 처리 오류: {img_error}")
                continue
                
        except requests.exceptions.Timeout:
            print(f"⏰ {chart_name}: 타임아웃 (시도 {attempt + 1})")
        except requests.exceptions.RequestException as e:
            print(f"🌐 {chart_name}: 네트워크 오류 - {e}")
        except Exception as e:
            print(f"❌ {chart_name}: 예상치 못한 오류 - {e}")
    
    print(f"💥 {chart_name}: 모든 시도 실패")
    return None

def insert_tableau_images_openpyxl_safe(ws, start_row):
    """안전한 Tableau 이미지 삽입 (오류 처리 강화)"""
    
    image_urls = get_tableau_image_urls()
    chart_titles = {
        "state_food": "🗺️ 미국 주별 식품 지출",
        "food_trend": "📈 연도별 식품 지출 추이",
        "recall_heatmap": "🔥 리콜 원인별 히트맵",
        "recall_class": "📊 리콜 등급별 발생 건수"
    }
    
    current_row = start_row
    
    # URL 접근성 먼저 테스트
    print("🔍 Tableau URL 접근성 테스트 중...")
    url_results = test_tableau_urls()
    
    accessible_charts = [name for name, result in url_results.items() if result.get('accessible', False)]
    print(f"✅ 접근 가능한 차트: {accessible_charts}")
    
    # Tableau 차트 섹션 제목
    ws[f'A{current_row}'] = "📊 Tableau 시장 분석 차트"
    ws[f'A{current_row}'].font = Font(size=14, bold=True)
    ws[f'A{current_row}'].fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    
    current_row += 2
    
    success_count = 0
    
    for chart_name, title in chart_titles.items():
        try:
            # 차트 제목 추가
            ws[f'A{current_row}'] = title
            ws[f'A{current_row}'].font = Font(size=12, bold=True)
            current_row += 1
            
            # URL 접근 가능성 확인
            if chart_name not in accessible_charts:
                ws[f'A{current_row}'] = f"[{title} - URL 접근 불가]"
                ws[f'A{current_row}'].font = Font(italic=True, color="FF6600")
                current_row += 2
                continue
            
            # 이미지 다운로드
            url = image_urls[chart_name]
            temp_image_path = download_tableau_image_safe(url, chart_name)
            
            if temp_image_path and os.path.exists(temp_image_path):
                try:
                    # openpyxl 이미지 객체 생성
                    img = image.Image(temp_image_path)
                    
                    # 이미지 크기 조정
                    img.width = 400
                    img.height = 280
                    
                    # 이미지 삽입
                    ws.add_image(img, f'A{current_row}')
                    
                    # 임시 파일 정리
                    try:
                        os.remove(temp_image_path)
                    except:
                        pass
                    
                    print(f"✅ {title} Excel 삽입 완료")
                    success_count += 1
                    
                    # 다음 이미지를 위한 공간 확보
                    current_row += 18
                    
                except Exception as excel_error:
                    print(f"❌ {title} Excel 삽입 실패: {excel_error}")
                    ws[f'A{current_row}'] = f"[{title} - Excel 삽입 실패]"
                    ws[f'A{current_row}'].font = Font(italic=True, color="FF0000")
                    current_row += 2
                    
            else:
                # 이미지 다운로드 실패
                ws[f'A{current_row}'] = f"[{title} - 이미지 다운로드 실패]"
                ws[f'A{current_row}'].font = Font(italic=True, color="808080")
                current_row += 2
                
        except Exception as e:
            print(f"❌ {title} 전체 처리 실패: {e}")
            ws[f'A{current_row}'] = f"[{title} - 처리 실패: {str(e)}]"
            ws[f'A{current_row}'].font = Font(italic=True, color="FF0000")
            current_row += 2
    
    # 결과 요약
    total_charts = len(chart_titles)
    ws[f'A{current_row}'] = f"📈 차트 삽입 결과: {success_count}/{total_charts} 성공"
    ws[f'A{current_row}'].font = Font(size=11, bold=True, 
                                      color="008000" if success_count == total_charts else "FF6600")
    
    return success_count

# =============================================================================
# Excel 생성 함수들
# =============================================================================

def create_smart_excel_report():
    """스마트 Excel 리포트 생성 - 이미지 시도 후 실패하면 텍스트로 자동 대체"""
    

    try:
        # 스타일 적용
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "미국 수출 상품 기획안"
        
        # 눈금선 숨기기
        ws.sheet_view.showGridLines = False
        ws.print_options.gridLines = False

        # 기본 정보 수집
        current_date = datetime.now().strftime('%Y년 %m월 %d일')
        product_name = st.session_state.get("product_name", "제품명 없음")
        target_name = st.session_state.get("target_name", "타겟층 없음")
        background = st.session_state.get("background", "추진배경 없음")
        summary_content = st.session_state.get("summary_content", "분석 내용 없음")
        
        # 1. 제목 및 헤더
        ws['B1'] = f"{product_name} 상품 기획안_미국 시장 진출 분석"
        ws['B1'].font = Font(size=18, bold=True, color="FFFFFF")
        ws['B1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('B1:G1')
        
        ws['B2'] = f"작성일: {current_date}"
        ws['B2'].font = Font(size=11)
        ws['B2'].alignment = Alignment(horizontal='right')
        ws.merge_cells('B2:G2')
        
        # 2. 제품 정보 섹션
        current_row = 4
        ws[f'B{current_row}'] = "📋 제품 정보"
        ws[f'B{current_row}'].font = Font(size=14, bold=True)
        ws[f'B{current_row}'].fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        ws['B4'].alignment = Alignment(horizontal='left',vertical='center')
        ws.merge_cells('B4:C4')

        current_row += 2
        ws[f'B{current_row}'] = "제품명:"
        ws[f'B{current_row}'].font = Font(bold=True)
        ws[f'C{current_row}'] = product_name
        
        current_row += 1
        ws[f'B{current_row}'] = "타겟층:"
        ws[f'B{current_row}'].font = Font(bold=True)
        ws[f'C{current_row}'] = target_name
        
        # 3. 기획배경 섹션
        current_row += 3
        ws[f'B{current_row}'] = "📊 기획 의도"
        ws[f'B{current_row}'].font = Font(size=14, bold=True)
        ws[f'B{current_row}'].fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        ws['B10'].alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells('B10:C10')

        current_row += 2
        ws[f'B{current_row}'] = background[:1000]
        ws[f'B{current_row}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(f'B{current_row}:G{current_row+3}')
        
        # 4. AI 분석 결과 섹션
        current_row += 5
        ws[f'B{current_row}'] = "🤖 AI 분석 결과"
        ws[f'B{current_row}'].font = Font(size=14, bold=True)
        ws[f'B{current_row}'].fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        ws['B17'].alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells('B17:C17')
        
        current_row += 2
        ws[f'B{current_row}'] = summary_content[:1500]
        ws[f'B{current_row}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(f'B{current_row}:G{current_row+5}')
        
        # 5. 스마트 차트 섹션 (이미지 시도 후 실패하면 텍스트)
        current_row += 8
        chart_success_count = insert_smart_tableau_charts(ws, current_row)
        
        # 6. 기본 스타일 적용 (컬럼 너비와 행 높이만)  ← 여기에 추가
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 25

        ws.row_dimensions[1].height = 50
        ws.row_dimensions[4].height = 25
        ws.row_dimensions[10].height = 25
        ws.row_dimensions[17].height = 25
        ws.row_dimensions[19].height = 30
        ws.row_dimensions[20].height = 30
        ws.row_dimensions[21].height = 30
        ws.row_dimensions[22].height = 30
        ws.row_dimensions[23].height = 30
        ws.row_dimensions[24].height = 30
        ws.row_dimensions[27].height = 25

        # 7. 파일 저장 (프로젝트명 포함)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        # 선택된 프로젝트명 가져오기
        selected_project = st.session_state.get("project_selector", "새 프로젝트")
        if selected_project and selected_project != "새 프로젝트":
            # 파일명에 사용할 수 없는 문자 제거
            selected_project_name = "".join(c for c in selected_project if c.isalnum() or c in (' ', '-', '_')).strip()
            selected_project_name = selected_project_name.replace(' ', '_')
            filename = f"{selected_project_name}_상품기획안_{timestamp}.xlsx"
        else:
            filename = f"상품기획안_{timestamp}.xlsx"

        wb.save(filename)
        
        return True, filename, chart_success_count
        
    except Exception as e:
        return False, f"Excel 생성 중 오류: {str(e)}", 0

def insert_smart_tableau_charts(ws, start_row):
    """로컬 이미지 우선 삽입, 없으면 텍스트"""
    
    charts_config = {
        "state_food": "🗺️ 미국 주별 식품 지출",
        "food_trend": "📈 연도별 식품 지출 추이",
        "recall_heatmap": "🔥 리콜 원인별 히트맵",
        "recall_class": "📊 리콜 등급별 발생 건수"
    }
    
    current_row = start_row
    success_count = 0
    
    # 차트 섹션 제목
    ws[f'B{current_row}'] = "📊 미국 시장 분석 차트"
    ws[f'B{current_row}'].font = Font(size=14, bold=True)
    ws[f'B{current_row}'].fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    ws['B27'].alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells('B27:C27')
    current_row += 2
    
    # 먼저 이미지가 있는지 확인
    available_images = []
    for chart_name in charts_config.keys():
        image_path = f"./charts/{chart_name}.png"
        if os.path.exists(image_path):
            available_images.append(chart_name)

    print(f"🔍 사용 가능한 이미지: {available_images}")

    if available_images:
        chart_items = list(charts_config.items())
        
        for i, (chart_name, title) in enumerate(chart_items):
            # 2개씩 배치: 0,1번은 첫 번째 행, 2,3번은 두 번째 행
            row_offset = (i // 2) * 15  # 행 간격 (이미지 크기 고려)
            col_offset = (i % 2) * 3 + 1   # 열 간격
            
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(col_offset + 1) 
            
            chart_row = current_row + row_offset
            
            # 차트 제목
            ws[f'{col_letter}{chart_row}'] = title
            ws[f'{col_letter}{chart_row}'].font = Font(size=12, bold=True)
            
            # 이미지 파일 경로
            image_path = f"./charts/{chart_name}.png"
            
            if os.path.exists(image_path):
                try:
                    # 이미지 삽입 (제목 아래)
                    img = image.Image(image_path)
                    img.width = 400  # 크기 조정
                    img.height = 280
                    ws.add_image(img, f'{col_letter}{chart_row + 1}')
                    
                    success_count += 1
                    print(f"✅ {title} 이미지 삽입 완료")
                    
                except Exception as e:
                    print(f"❌ {title} 이미지 삽입 실패: {e}")
        
    else:
        # 이미지가 없으면 텍스트 안내
        chart_info = """
            미국 시장 분석을 위한 주요 차트들:

            🗺️ 미국 주별 식품 지출 현황
            📈 연도별 식품 지출 추이  
            🔥 리콜 원인별 히트맵
            📊 리콜 등급별 발생 건수

            ※ 차트 이미지는 'Tableau 차트 다운로드' 버튼으로 다운로드 후 생성하세요.
                    """
        
        ws[f'B{current_row}'] = chart_info.strip()
        ws[f'B{current_row}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws[f'B{current_row}'].fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
        ws.merge_cells(f'B{current_row}:G{current_row+10}')
    
    return success_count

def add_single_excel_button():
    """Excel 리포트 생성 버튼"""
    
    # 필수 데이터 체크
    required_fields = ["product_name", "target_name", "background"]
    has_required_data = all(st.session_state.get(field, "") for field in required_fields)
    
    if not has_required_data:
        st.warning("⚠️ 제품명, 타겟층, 기획의도를 모두 입력해주세요.")
        return
    
    # 처리 상태 체크
    is_processing = st.session_state.get("ai_processing", False)
    if is_processing:
        st.info("🔄 AI 분석 처리 중입니다...")
        return
    
    # 이미지 상태 확인
    chart_files = [
        "./charts/state_food.png",
        "./charts/food_trend.png",
        "./charts/recall_heatmap.png", 
        "./charts/recall_class.png"
    ]
    
    available_images = [f for f in chart_files if os.path.exists(f)]
    
    # 상태 표시
    if available_images:
        st.success(f"📊 {len(available_images)}/4개 차트 이미지 준비됨")
    else:
        st.info("📝 차트 이미지 없음 (다운로드 후 Excel 생성 권장)")
    
    # Excel 생성 버튼
    if st.button("📊 Excel 리포트 생성", use_container_width=True):
        with st.spinner("📝 Excel 리포트 생성 중..."):
            try:
                success, result, chart_count = create_smart_excel_report()
                
                if success:
                    if chart_count > 0:
                        st.success(f"✅ Excel 생성 완료! ({chart_count}개 차트 이미지 포함)")
                    else:
                        st.success("✅ Excel 생성 완료! (텍스트 기반)")
                        st.info("💡 차트 이미지를 포함하려면 먼저 'Tableau 차트 다운로드'를 실행하세요.")
                    
                    # 다운로드 버튼
                    with open(result, "rb") as file:
                        st.download_button(
                            label="📥 Excel 파일 다운로드",
                            data=file.read(),
                            file_name=result,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                    
                    # 임시 파일 정리
                    try:
                        os.remove(result)
                    except:
                        pass
                else:
                    st.error(f"❌ {result}")
                    
            except Exception as e:
                st.error(f"❌ Excel 생성 실패: {e}")

# def apply_excel_styling(ws):
#     """Excel 스타일 적용"""
    
#     # 전체 폰트 설정
#     for row in ws.iter_rows():
#         for cell in row:
#             if cell.value:
#                 cell.font = Font(name='맑은 고딕', size=10)
    
#     # 컬럼 너비 조정
#     ws.column_dimensions['A'].width = 20
#     ws.column_dimensions['B'].width = 30
#     ws.column_dimensions['C'].width = 25
#     ws.column_dimensions['D'].width = 25
#     ws.column_dimensions['E'].width = 25
#     ws.column_dimensions['F'].width = 25
    
#     # 행 높이 조정
#     ws.row_dimensions[1].height = 30
#     ws.row_dimensions[4].height = 25

# =============================================================================
# 레거시 함수 (호환성 유지)
# =============================================================================

def add_excel_export_button():
    """레거시 함수 - 호환성을 위해 유지"""
    add_single_excel_button()

def create_excel_with_tableau():
    """레거시 함수 - 호환성을 위해 유지"""
    return create_smart_excel_report()